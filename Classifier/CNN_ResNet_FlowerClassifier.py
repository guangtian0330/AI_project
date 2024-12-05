# -*- coding: utf-8 -*-
"""
CPSC-4117EL: CNN using Transfer Learning (from ResNet18)

A basic example that outlines how you can use a webcam to collect data,
train a simple CNN using <transfer learning>, and test the model using
real-time camera input.

Flow of the program:

1. Collect Data using Webcam:
- press 'c' key to capture images as input, by default 10 images for each classes

2. Train our CNN using Transfer Learning (from ResNet - residual networks):
- we initialize the ResNet-18 model (18 layers) with weights pre-trained on the ImageNet dataset.
We freeze all layers (i.e., you don't update their weights during training)
except for the final fully connected layer, which will adapt to our
three specific classes. This approach is effective because the model leverages
the general features learned from the ImageNet dataset and only
fine-tunes the decision-making (final layer) for your specific classes.

https://pytorch.org/vision/main/models/resnet.html

This does not mean we won't be training on our own dataset.
Instead, we'll be starting with a model that already has learned
features from ImageNet, which can give us a significant boost,
especially when our dataset is small. This method is called "Transfer Learning"

https://deeplearning.cms.waikato.ac.nz/user-guide/class-maps/IMAGENET/
https://debuggercafe.com/implementing-resnet18-in-pytorch-from-scratch/

ResNet-18: 1 (initial conv) + 2x2 (first set) + 2x2 (second set) +
2x2 (third set) + 2x2 (fourth set) + 1 (fully connected) = 18 layers

-Details of Layers-
Initial Convolution + Pooling
The network starts with:
1 Convolutional Layer: A 7x7 kernel with stride 2 for down-sampling.
Followed by Batch Normalization, ReLU, and a MaxPooling layer (3x3 kernel, stride 2).
Total layers contributing to the count: 1 layer (only the convolution layer is counted).

First Residual Block (2 × 2 Layers): Input → Conv1 → BN → ReLU → Conv2 → BN → Add (skip connection) → ReLU → Output.
Contains 2 residual units:
Each unit has 2 convolutional layers (3x3 kernels with padding).
Each layer is followed by batch normalization and ReLU.
Total convolutional layers: 2 × 2 = 4 layers.

Second Residual Block (2 × 2 Layers)
Similar to the first block but down-samples the feature map size (stride 2 in the first convolution of the block).
Total convolutional layers: 2 × 2 = 4 layers.

Third Residual Block (2 × 2 Layers)
Same structure as previous blocks with down-sampling in the first convolution of the block.
Total convolutional layers: 2 × 2 = 4 layers.

Fourth Residual Block (2 × 2 Layers)
Same structure again, with down-sampling in the first convolution of the block.
Total convolutional layers: 2 × 2 = 4 layers.

Final Fully Connected Layer
After global average pooling, a fully connected layer outputs predictions.
Total layers contributing to the count: 1 layer.

3. Test the Model with Webcam Input:
- point your webcam to the image to see the classification in real-time.
- press 'q' key to quit
"""
import cv2
import os
import torch
import torch.nn as nn
import torch.optim as optim
import torchvision.transforms as transforms
from torchvision.models import resnet34
from torch.utils.data import DataLoader, TensorDataset
from openpyxl import Workbook, load_workbook
import pandas as pd
from sklearn.metrics import accuracy_score, f1_score, confusion_matrix

class FlowerClassifier:
    def __init__(self, class_names=["Daisy", "Lily", "Jasmine"], num_samples=20, num_epochs=10, batch_size=4, lr=0.001,
                 output_file="predictions.xlsx"):
        self.class_names = class_names
        self.num_samples = num_samples
        self.num_epochs = num_epochs
        self.batch_size = batch_size
        self.lr = lr
        self.model_path = "model_state.pth"
        self.output_file = output_file
        if not os.path.exists(self.output_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Prediction", "True Result", "Match"])  # Add header
            wb.save(self.output_file)
        else:
            self.wb = load_workbook(self.output_file)
        # Check if 'data' directory exists. If not, create directories for three classes.
        if not os.path.exists('data'):
            for class_name in self.class_names:
                os.makedirs(f'data/{class_name}')

        # Setup device for PyTorch (use CUDA if available).
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

        # Initialize model
        self.model = self._initialize_model()
        # Define the loss function and optimizer.
        self.criterion = nn.CrossEntropyLoss()
        self.optimizer = optim.Adam(self.model.fc.parameters(), lr=self.lr)

        def preprocess_with_denoise(frame):
            # 转为灰度图
            gray_image = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            # 应用高斯模糊
            blurred_image = cv2.GaussianBlur(gray_image, (5, 5), 0)
            # 转回彩色
            processed_image = cv2.cvtColor(blurred_image, cv2.COLOR_GRAY2BGR)
            return processed_image

        class CustomTransform:
            def __call__(self, frame):
                # 调用自定义的 OpenCV 处理逻辑
                processed_frame = preprocess_with_denoise(frame)
                return processed_frame

        # Define the transformations for preprocessing the images.
        self.data_transforms = transforms.Compose([
            CustomTransform(),  # Custom preprocessing step
            transforms.ToPILImage(),  # Convert image to PIL Image format.
            transforms.Resize((256, 256)),  # Resize to fit ResNet18's input size.
            # Apply random augmentations to increase dataset variety.
            transforms.RandomHorizontalFlip(),
            transforms.RandomRotation(15),
            transforms.ColorJitter(brightness=0.2, contrast=0.2, saturation=0.2),
            transforms.ToTensor(),
            # Normalize according to values suitable for ResNet18.
            transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
        ])

    def _initialize_model(self):
        # Load the pre-trained ResNet18 model.
        model = resnet34(pretrained=True)
        # Freeze all layers in the model. We'll only train the final layer.
        for param in model.parameters():
            param.requires_grad = False
        # Modify the final fully connected layer to classify 3 classes.
        model.fc = nn.Linear(model.fc.in_features, len(self.class_names))
        model = model.to(self.device)
        return model

    """
    Collect data for all 3 different types of flowers.
    """
    def collect_data(self):
        """Collects data from the webcam for each class."""
        cap = cv2.VideoCapture(0)
        for class_name in self.class_names:
            count = 0
            print(f'Collecting {class_name}')
            while count < self.num_samples:
                ret, frame = cap.read()
                if not ret:
                    break
                cv2.imshow('Collect Data', frame)
                if cv2.waitKey(1) & 0xFF == ord('c'):
                    cv2.imwrite(f'data/{class_name}/{class_name}_{count}.jpg', frame)
                    print(f'Sample {count + 1} captured')
                    count += 1
            print(f"Finished collecting data for {class_name}")
        cap.release()
        cv2.destroyAllWindows()

    def load_data(self):
        # Read and preprocess images from each class directory.
        data, labels = [], []
        for class_name in self.class_names:
            for filename in os.listdir(f'data/{class_name}'):
                img = cv2.imread(f'data/{class_name}/{filename}')
                img = self.data_transforms(img)
                data.append(img)
                # Get the index of the class and append it to the labels.
                labels.append(self.class_names.index(class_name))

        # Convert data and labels to PyTorch tensors.
        data = torch.stack(data)
        labels = torch.tensor(labels)

        # Create a PyTorch dataset and data loader.
        dataset = TensorDataset(data, labels)
        dataloader = DataLoader(dataset, batch_size=self.batch_size, shuffle=True)
        return dataloader

    def train(self, dataloader):
        # Training loop for the model.
        self.model.train()
        for epoch in range(self.num_epochs):
            for inputs, labels in dataloader:
                inputs, labels = inputs.to(self.device), labels.to(self.device)
                self.optimizer.zero_grad()  # Zero out any gradient from the previous step.
                outputs = self.model(inputs)
                loss = self.criterion(outputs, labels)
                loss.backward()  # Compute the gradients.
                self.optimizer.step()  # Update the weights.
            print(f"Epoch {epoch + 1}/{self.num_epochs}, Loss: {loss.item():.4f}")

    def save_model(self):
        # Save the model's state dictionary
        torch.save(self.model.state_dict(), self.model_path)
        print(f"Model saved to {self.model_path}")

    def load_model(self):
        """Loads the model's state dictionary."""
        self.model.load_state_dict(torch.load(self.model_path))
        self.model.eval()  # Set the model to evaluation mode if using for inference
        print(f"Model loaded from {self.model_path}")

    def predict(self):
        """Performs real-time predictions using the webcam feed."""
        cap = cv2.VideoCapture(0)
        true_label = None
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            # Preprocess the frame and get predictions.
            # Transform the image frame: increases the dimensions of the tensor
            # by adding an extra dimension at the 0th position to create
            # a batch of one image for ResNet18
            input_img = self.data_transforms(frame).unsqueeze(0).to(self.device)
            with torch.no_grad():  #  It doesn't need to keep track of gradients.
                output = self.model(input_img)
                _, pred = torch.max(output, 1)  # get the predicted class

            # Map numeric predictions back to class names.
            label = self.class_names[pred.item()]
            # Display the predictions on the frame.
            cv2.putText(frame, label, (50, 70), cv2.FONT_HERSHEY_SIMPLEX, 2, (255, 0, 0), 2)
            cv2.imshow('Predictions', frame)

            # Exit loop on pressing 'q' key.
            key = cv2.waitKey(1) & 0xFF
            if key == ord('d'):
                true_label = "Daisy"
                print("THE Current picture is a daisy")
            elif key == ord('l'):
                true_label = "Lily"
                print("THE Current picture is a lily")
            elif key == ord('j'):
                true_label = "Jasmine"
                print("THE Current picture is a jasmine")
            elif key == ord('e'):
                true_label = None
                print("Current prediction is finished, waiting for the next one.")
            elif key == ord('q'):
                break  # Exit the loop on 'q'
            if true_label is not None:
                # Compare prediction with true label
                match = 1 if label == true_label else 0
                # Write the results to Excel
                self.write_to_excel(label, true_label, match)
        cap.release()
        cv2.destroyAllWindows()

    def write_to_excel(self, prediction, true_label, match):
        wb = load_workbook(self.output_file)
        ws = wb.active
        ws.append([prediction, true_label, match])  # Append the row
        wb.save(self.output_file)

    def calculate_metrics(self):
        data = pd.read_excel(self.output_file)
        required_columns = ["Prediction", "True Result"]
        if not all(col in data.columns for col in required_columns):
            raise ValueError(f"The Excel file must contain the columns: {required_columns}")
        predictions = data["Prediction"]
        true_labels = data["True Result"]
        # Calculate the accuracy
        accuracy = accuracy_score(true_labels, predictions)
        # Calculate F1 Score
        f1 = f1_score(true_labels, predictions, average="macro")
        # Generate the confusion matrix
        conf_matrix = confusion_matrix(true_labels, predictions, labels=true_labels.unique())
        # Print the evaluation results.
        print(f"Accuracy: {accuracy:.4f}")
        print(f"F1 Score (Macro): {f1:.4f}")
        print("Confusion Matrix:")
        print(conf_matrix)
        return accuracy, f1, conf_matrix

def main():
    # Create an instance of the FlowerClassifier class
    flower_classifier = FlowerClassifier()
    # If the model is not created, then creat a model and train it.
    if not os.path.exists(flower_classifier.model_path):
        # Collect data from webcam
        flower_classifier.collect_data()
        # Load data
        dataloader = flower_classifier.load_data()
        # Train and save the model
        flower_classifier.train(dataloader)
        flower_classifier.save_model()
    # Load the trained model for predictions
    flower_classifier.load_model()
    # Run real-time predictions using webcam
    flower_classifier.predict()
    accuracy, f1, conf_matrix = flower_classifier.calculate_metrics()


if __name__ == "__main__":
    main()